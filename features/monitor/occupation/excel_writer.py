from __future__ import annotations

import io
import json
import logging
import os
from datetime import date, timedelta
from typing import List

import pandas as pd

from config import APP_ROOT, OCCUPATION_EXCEL_PATHS
from features.monitor.occupation.smb_excel import read_excel_bytes, write_excel_bytes

_log = logging.getLogger(__name__)

PATHS_JSON = os.path.join(APP_ROOT, "occupation_paths.json")

_MONTH_SHEET = {
    1: "JANUARY", 2: "FEBRUARY", 3: "MARCH", 4: "APRIL",
    5: "MAY", 6: "JUNE", 7: "JULY", 8: "AUGUST",
    9: "SEPTEMBER", 10: "OCTOBER", 11: "NOVEMBER", 12: "DECEMBER",
}

_POS_COL = {1: "G", 2: "L"}


def _load_paths() -> dict:
    """Return path map from occupation_paths.json if present, else config defaults."""
    if os.path.exists(PATHS_JSON):
        try:
            with open(PATHS_JSON, encoding="utf-8") as fh:
                return json.load(fh)
        except Exception as exc:
            _log.warning("Could not read %s: %s — using config defaults", PATHS_JSON, exc)
    return dict(OCCUPATION_EXCEL_PATHS)


def save_paths(paths: dict) -> None:
    """Persist the path map to occupation_paths.json after basic validation."""
    errors = []
    for machine_id, path in paths.items():
        if not path or not path.strip():
            errors.append(f"{machine_id}: path is empty")
        elif not path.strip().lower().endswith(".xlsm"):
            errors.append(f"{machine_id}: path must end with .xlsm")
    if errors:
        raise ValueError("; ".join(errors))
    with open(PATHS_JSON, "w", encoding="utf-8") as fh:
        json.dump(paths, fh, indent=2)
    _log.info("Occupation paths saved to %s", PATHS_JSON)


def detect_breaks(df: pd.DataFrame, target_date: date) -> str:
    """Return a break string for target_date from a parsed log DataFrame.

    Each continuous zero-speed interval becomes one 'Break:HH:MM-HH:MM' line.
    Multiple breaks are joined with newline. Returns '' when speed is never 0.
    """
    df = df.copy()
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    day_df = df[df["timestamp"].dt.date == target_date].sort_values("timestamp")
    if day_df.empty:
        return ""

    speed = pd.to_numeric(day_df["speed"], errors="coerce").fillna(0)
    breaks: list[str] = []
    in_break = False
    start_ts = None

    for ts, is_zero in zip(day_df["timestamp"], speed == 0):
        if is_zero and not in_break:
            in_break = True
            start_ts = ts
        elif not is_zero and in_break:
            in_break = False
            breaks.append(f"Break:{start_ts.strftime('%H:%M')}-{ts.strftime('%H:%M')}")

    if in_break and start_ts is not None:
        end_ts = day_df["timestamp"].iloc[-1]
        breaks.append(f"Break:{start_ts.strftime('%H:%M')}-{end_ts.strftime('%H:%M')}")

    return "\n".join(breaks)


def dates_in_range(start: date, end: date) -> List[date]:
    """Return all dates from start to end inclusive."""
    result = []
    current = start
    while current <= end:
        result.append(current)
        current += timedelta(days=1)
    return result


def fill_occupation(
    machine_id: str,
    position: int,
    dates: List[date],
    df: pd.DataFrame,
) -> str:
    """Read the .xlsm, fill break data for each date/position, write back.

    Returns a human-readable status string (success or error).
    """
    paths = _load_paths()
    rel_path = paths.get(machine_id)
    if not rel_path:
        return f"No Excel path configured for {machine_id}."

    col_letter = _POS_COL.get(position)
    if not col_letter:
        return f"No column configured for position {position}."

    # Filter DataFrame to the given position
    pos_df = df[df["position"] == position] if "position" in df.columns else df

    try:
        raw_bytes = read_excel_bytes(rel_path)
    except Exception as exc:
        _log.error("Failed to read Excel %s: %s", rel_path, exc)
        return f"Error reading file: {exc}"

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), keep_vba=True)
    except ImportError:
        return "openpyxl is not installed. Run: pip install openpyxl"
    except Exception as exc:
        _log.error("openpyxl failed to open %s: %s", rel_path, exc)
        return f"Error opening workbook: {exc}"

    filled = 0
    skipped = []
    for d in dates:
        sheet_name = _MONTH_SHEET.get(d.month)
        if not sheet_name or sheet_name not in wb.sheetnames:
            skipped.append(f"{d.isoformat()} (sheet {sheet_name} not found)")
            continue

        ws = wb[sheet_name]
        date_str = d.strftime("%d/%m/%Y")
        target_row = None

        for row in ws.iter_rows(min_col=2, max_col=2):
            cell = row[0]
            cell_val = cell.value
            if cell_val is None:
                continue
            if hasattr(cell_val, "strftime"):
                formatted = cell_val.strftime("%d/%m/%Y")
            else:
                formatted = str(cell_val).strip()
            if formatted == date_str:
                target_row = cell.row
                break

        if target_row is None:
            skipped.append(f"{d.isoformat()} (date not found in sheet)")
            continue

        break_str = detect_breaks(pos_df, d)
        ws[f"{col_letter}{target_row}"] = break_str
        filled += 1
        _log.info("Wrote break data for %s pos%d @ %s row %d", machine_id, position, date_str, target_row)

    buf = io.BytesIO()
    try:
        wb.save(buf)
    except Exception as exc:
        _log.error("openpyxl save failed: %s", exc)
        return f"Error saving workbook: {exc}"

    try:
        write_excel_bytes(rel_path, buf.getvalue())
    except Exception as exc:
        _log.error("Failed to write Excel %s: %s", rel_path, exc)
        return f"Error writing file: {exc}"

    parts = [f"{filled} date(s) written successfully."]
    if skipped:
        parts.append("Skipped: " + "; ".join(skipped))
    return " ".join(parts)
